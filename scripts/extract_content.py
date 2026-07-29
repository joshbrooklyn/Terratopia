"""
One-off script: extract rarity-1 Techs and Items from the spreadsheet
and write them to GameData/Techs/techs.json and GameData/Items/items.json.

Run from the repo root:
    python scripts/extract_content.py
"""

import json
import os
import re
import openpyxl

SPREADSHEET = os.path.join("Reference", "Terratopia Adventurer Combat Resources.xlsx")
TECH_SHEETS = {
    "Black Mage Techs": "BlackMage",
    "White Mage Techs": "WhiteMage",
    "Rogue Techs":      "Rogue",
    "Fighter Techs":    "Fighter",
}
RARITY_1 = "1. Low Common"


# ── Helpers ─────────────────────────────────────────────────────────────────

def to_id(name: str) -> str:
    """'Cure Greatest Wounds' → 'cure_greatest_wounds'"""
    s = name.lower()
    s = re.sub(r"[^a-z0-9\s]", "", s)
    s = re.sub(r"\s+", "_", s.strip())
    return s


def status_id(raw: str) -> str:
    """'P Dwn 2' → 'PDwn2', 'HP-3' → 'HpMinus3', 'HP+ 3' → 'HpPlus3'"""
    s = raw.strip()
    # Expand + and - into words so they survive the word-split
    s = re.sub(r"\+", " Plus ", s)
    s = re.sub(r"-",  " Minus ", s)
    return "".join(w.capitalize() for w in s.split())


def parse_statuses(cell) -> list[str]:
    if not cell:
        return []
    return [status_id(s.strip()) for s in str(cell).split(",") if s.strip()]


def parse_traits(cell) -> list[str]:
    """Normalize trait strings to ActionFlag names."""
    trait_map = {
        "ignore def":       "IgnoresDefense",
        "no evade":         "Unevadeable",
        "no cover":         "Unredirectable",
        "fixed dmg":        "FixedValue",
        "fixed pwr":        "FixedPower",
        "affects tp":       "AffectsTp",
        "out of combat":    "UsableOutsideCombat",
        "healing":          "Healing",
        "targets self":     "TargetsSelf",
        "target filtering": "TargetFiltering",
    }
    if not cell:
        return []
    result = []
    for raw in str(cell).split(","):
        key = raw.strip().lower()
        result.append(trait_map.get(key, raw.strip()))
    return result


def parse_keywords(cell) -> list[str]:
    if not cell:
        return []
    return [k.strip() for k in str(cell).split(",") if k.strip()]


def parse_element(cell) -> str | None:
    if not cell or str(cell).strip().lower() in ("none", ""):
        return None
    raw = str(cell).strip()
    element_map = {"lit": "Lightning", "fire": "Fire", "ice": "Ice", "void": "Void"}
    return element_map.get(raw.lower(), raw)


def parse_targeting(targeting_cell, valid_targets_cell, traits: list[str]) -> str:
    if "TargetsSelf" in traits:
        return "Self"
    t = str(targeting_cell).strip() if targeting_cell else "Choose"
    v = str(valid_targets_cell).strip() if valid_targets_cell else "Opponents"
    mapping = {
        "Choose":               "ChooseWithReplacement",
        "Choose No Replacement":"ChooseWithoutReplacement",
        "Random":               "RandomWithReplacement",
        "Random No Replacement":"RandomWithoutReplacement",
    }
    if t == "All":
        if v == "Opponents": return "AllEnemies"
        if v == "Allies":    return "AllAllies"
        return "AllEntities"
    return mapping.get(t, "ChooseWithReplacement")


def parse_num_attacks(cell) -> int:
    if cell is None:
        return 1
    if isinstance(cell, str) and cell.strip().lower() == "all":
        return 1
    try:
        v = int(float(cell))
        return max(v, 1)
    except (ValueError, TypeError):
        return 1


def derive_effect_type(power, traits: list[str]) -> str | None:
    if "Healing" in traits:
        return "Heal"
    try:
        if float(power) > 0:
            return "Damage"
    except (TypeError, ValueError):
        pass
    return None


# ── Tech extraction ──────────────────────────────────────────────────────────

def extract_techs(wb) -> list[dict]:
    techs = []
    for sheet_name, job_class in TECH_SHEETS.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:
            if row[3] != RARITY_1:
                continue
            name        = row[0]
            tp_cost     = int(row[1]) if row[1] is not None else 0
            tier        = int(row[2]) if row[2] is not None else 1
            description = str(row[4]) if row[4] else ""
            power       = row[5] if row[5] is not None else 0.0
            keywords    = parse_keywords(row[6])
            traits      = parse_traits(row[7])
            num_attacks = parse_num_attacks(row[8])
            targeting   = parse_targeting(row[9], row[10], traits)
            element     = parse_element(row[12])
            t_statuses  = parse_statuses(row[13])
            u_statuses  = parse_statuses(row[14])
            effect_type = derive_effect_type(power, traits)

            # Manual correction: Blinding Powder's user evasion gain is
            # missing from the User Status column — add it here.
            if name == "Blinding Powder" and not u_statuses:
                u_statuses = ["EvasionPlus40Pct"]

            # Remove "TargetsSelf" from the traits list — it drove targeting,
            # but isn't a flag the engine understands.
            traits = [t for t in traits if t != "TargetsSelf"]

            techs.append({
                "techId":        to_id(name),
                "name":          name,
                "jobClass":      job_class,
                "tpCost":        tp_cost,
                "tier":          tier,
                "rarity":        1,
                "description":   description,
                "powerModifier": float(power),
                "numAttacks":    num_attacks,
                "keywords":      keywords,
                "traits":        traits,
                "targetingType": targeting,
                "element":       element,
                "effectType":    effect_type,
                "targetStatuses":t_statuses,
                "userStatuses":  u_statuses,
            })
    return techs


# ── Item extraction ──────────────────────────────────────────────────────────

def extract_items(wb) -> list[dict]:
    items = []
    ws = wb["Items"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:
        if row[1] != RARITY_1:
            continue
        name        = row[0]
        description = str(row[2]) if row[2] else ""
        power       = row[3] if row[3] is not None else 0.0
        traits      = parse_traits(row[5])
        num_attacks = parse_num_attacks(row[6])
        targeting   = parse_targeting(row[7], row[8], traits)
        element     = parse_element(row[10])
        t_statuses  = parse_statuses(row[11])
        u_statuses  = parse_statuses(row[12])
        effect_type = derive_effect_type(power, traits)

        traits = [t for t in traits if t != "TargetsSelf"]

        items.append({
            "itemId":        to_id(name),
            "name":          name,
            "rarity":        1,
            "description":   description,
            "powerModifier": float(power),
            "numAttacks":    num_attacks,
            "traits":        traits,
            "targetingType": targeting,
            "element":       element,
            "effectType":    effect_type,
            "targetStatuses":t_statuses,
            "userStatuses":  u_statuses,
        })
    return items


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.load_workbook(SPREADSHEET, read_only=True, data_only=True)

    techs = extract_techs(wb)
    items = extract_items(wb)

    game_data_root = os.environ["TerratopiaGameDataPath"]

    os.makedirs(os.path.join(game_data_root, "Techs"), exist_ok=True)
    os.makedirs(os.path.join(game_data_root, "Items"), exist_ok=True)

    tech_path = os.path.join(game_data_root, "Techs", "techs.json")
    item_path = os.path.join(game_data_root, "Items", "items.json")

    with open(tech_path, "w", encoding="utf-8") as f:
        json.dump(techs, f, indent=2, ensure_ascii=False)

    with open(item_path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)

    print(f"Wrote {len(techs)} techs -> {tech_path}")
    print(f"Wrote {len(items)} items -> {item_path}")


if __name__ == "__main__":
    main()
